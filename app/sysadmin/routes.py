"""
System Admin Blueprint — Full Database Management Panel
Only accessible to users with role = SYSTEM_ADMIN
"""
from flask import (
    Blueprint, render_template, request, jsonify,
    redirect, url_for, flash, make_response, abort
)
from flask_login import login_required, current_user
from sqlalchemy import inspect as sa_inspect, text
from app.extensions import db
from app.models import (
    Roles, User, Student, Teacher, Department, Class,
    Subject, Attendance, Marks, Notice, Grievance,
    Certificate, Assignment, Notification, Status
)
import io, json
from datetime import datetime

sysadmin_bp = Blueprint('sysadmin', __name__, url_prefix='/sysadmin')

# ── Guard decorator ────────────────────────────────────────────────────────────
def sysadmin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != Roles.SYSTEM_ADMIN:
            abort(403)
        return f(*args, **kwargs)
    return decorated

# ── Table Registry: maps URL-slug → model + metadata ──────────────────────────
TABLE_REGISTRY = {
    'users':        {'model': User,         'label': 'Users',           'icon': '👤', 'color': '#6c63ff'},
    'students':     {'model': Student,      'label': 'Students',        'icon': '🎓', 'color': '#0ea5e9'},
    'teachers':     {'model': Teacher,      'label': 'Teachers',        'icon': '👩‍🏫', 'color': '#10b981'},
    'departments':  {'model': Department,   'label': 'Departments',     'icon': '🏢', 'color': '#f59e0b'},
    'classes':      {'model': Class,        'label': 'Classes',         'icon': '🏫', 'color': '#ec4899'},
    'subjects':     {'model': Subject,      'label': 'Subjects',        'icon': '📚', 'color': '#8b5cf6'},
    'attendance':   {'model': Attendance,   'label': 'Attendance',      'icon': '✅', 'color': '#14b8a6'},
    'marks':        {'model': Marks,        'label': 'Marks',           'icon': '📊', 'color': '#f97316'},
    'notices':      {'model': Notice,       'label': 'Notices',         'icon': '📢', 'color': '#06b6d4'},
    'grievances':   {'model': Grievance,    'label': 'Grievances',      'icon': '📋', 'color': '#ef4444'},
    'certificates': {'model': Certificate, 'label': 'Certificates',    'icon': '🎖️', 'color': '#a855f7'},
    'assignments':  {'model': Assignment,   'label': 'Assignments',     'icon': '📝', 'color': '#84cc16'},
    'notifications':{'model': Notification,'label': 'Notifications',   'icon': '🔔', 'color': '#eab308'},
}

def _get_columns(model):
    """Return column definitions for a model using SQLAlchemy inspection."""
    mapper = sa_inspect(model)
    cols = []
    for attr in mapper.column_attrs:
        col = attr.columns[0]
        cols.append({
            'name': attr.key,
            'type': str(col.type.__class__.__name__),
            'nullable': col.nullable,
            'primary_key': col.primary_key,
        })
    return cols

def _record_to_dict(record, columns):
    """Safely serialize a SQLAlchemy record to a plain dict."""
    d = {}
    for col in columns:
        val = getattr(record, col['name'], None)
        if isinstance(val, datetime):
            val = val.strftime('%Y-%m-%d %H:%M:%S')
        elif hasattr(val, 'value'):          # Enum
            val = val.value
        elif val is not None:
            val = str(val)
        d[col['name']] = val
    return d

# ── Dashboard ──────────────────────────────────────────────────────────────────
@sysadmin_bp.route('/')
@login_required
@sysadmin_required
def index():
    table_stats = []
    for slug, meta in TABLE_REGISTRY.items():
        try:
            count = db.session.query(meta['model']).count()
        except Exception:
            db.session.rollback()
            count = '?'
        table_stats.append({'slug': slug, 'label': meta['label'],
                             'icon': meta['icon'], 'color': meta['color'],
                             'count': count})

    # DB size — works on PostgreSQL only
    db_size = 'N/A'
    try:
        result = db.session.execute(
            text("SELECT pg_size_pretty(pg_database_size(current_database()))")
        ).scalar()
        if result:
            db_size = result
    except Exception:
        db.session.rollback()

    total_records = sum(t['count'] for t in table_stats if isinstance(t['count'], int))

    return render_template('sysadmin/index.html',
        table_stats=table_stats,
        db_size=db_size,
        total_records=total_records,
        active_page='dashboard'
    )

# ── Table Browser ──────────────────────────────────────────────────────────────
@sysadmin_bp.route('/table/<slug>')
@login_required
@sysadmin_required
def table_view(slug):
    if slug not in TABLE_REGISTRY:
        abort(404)
    meta    = TABLE_REGISTRY[slug]
    model   = meta['model']
    columns = _get_columns(model)

    page        = request.args.get('page', 1, type=int)
    per_page    = request.args.get('per_page', 50, type=int)
    search      = request.args.get('q', '').strip()
    sort_col    = request.args.get('sort', 'id')
    sort_dir    = request.args.get('dir', 'asc')

    query = model.query

    # Search across text columns
    if search:
        from sqlalchemy import or_, cast, String
        col_names = [c['name'] for c in columns if c['type'] in ('String', 'Text', 'VARCHAR')]
        filters = []
        for cn in col_names:
            col_attr = getattr(model, cn, None)
            if col_attr is not None:
                filters.append(cast(col_attr, String).ilike(f'%{search}%'))
        if filters:
            query = query.filter(or_(*filters))

    # Sort
    sort_attr = getattr(model, sort_col, None)
    if sort_attr is not None:
        query = query.order_by(sort_attr.desc() if sort_dir == 'desc' else sort_attr.asc())

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    records    = [_record_to_dict(r, columns) for r in pagination.items]

    return render_template('sysadmin/table.html',
        slug=slug, meta=meta, columns=columns,
        records=records, pagination=pagination,
        search=search, sort_col=sort_col, sort_dir=sort_dir,
        per_page=per_page,
        table_stats=[{'slug': s, 'label': m['label'], 'icon': m['icon'],
                      'color': m['color'], 'count': '...'}
                     for s, m in TABLE_REGISTRY.items()],
        active_page=slug
    )

# ── Export to Excel ────────────────────────────────────────────────────────────
@sysadmin_bp.route('/table/<slug>/export')
@login_required
@sysadmin_required
def export_excel(slug):
    if slug not in TABLE_REGISTRY:
        abort(404)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxl not installed. Run: pip install openpyxl', 'danger')
        return redirect(url_for('sysadmin.table_view', slug=slug))

    meta    = TABLE_REGISTRY[slug]
    model   = meta['model']
    columns = _get_columns(model)

    search  = request.args.get('q', '').strip()
    query   = model.query
    if search:
        from sqlalchemy import or_, cast, String
        col_names = [c['name'] for c in columns if c['type'] in ('String', 'Text', 'VARCHAR')]
        filters = [cast(getattr(model, cn), String).ilike(f'%{search}%')
                   for cn in col_names if getattr(model, cn, None) is not None]
        if filters:
            query = query.filter(or_(*filters))

    records = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = meta['label']

    # Header style
    header_fill = PatternFill('solid', fgColor='1E293B')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [c['name'].replace('_', ' ').title() for c in columns]
    ws.append(headers)
    for i, cell in enumerate(ws[1], 1):
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = max(15, len(headers[i-1]) + 4)

    # Data rows
    alt_fill = PatternFill('solid', fgColor='F8FAFC')
    for row_idx, record in enumerate(records, 2):
        d = _record_to_dict(record, columns)
        row_data = [d.get(c['name'], '') for c in columns]
        ws.append(row_data)
        for cell in ws[row_idx]:
            cell.border    = border
            cell.alignment = Alignment(vertical='center')
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"{slug}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp  = make_response(buf.read())
    resp.headers['Content-Type']        = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    return resp

# ── Export to CSV ──────────────────────────────────────────────────────────────
@sysadmin_bp.route('/table/<slug>/export-csv')
@login_required
@sysadmin_required
def export_csv(slug):
    if slug not in TABLE_REGISTRY:
        abort(404)
    import csv
    meta    = TABLE_REGISTRY[slug]
    model   = meta['model']
    columns = _get_columns(model)
    records = model.query.all()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([c['name'] for c in columns])
    for r in records:
        d = _record_to_dict(r, columns)
        writer.writerow([d.get(c['name'], '') for c in columns])

    fname = f"{slug}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    resp  = make_response(buf.getvalue())
    resp.headers['Content-Type']        = 'text/csv'
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    return resp

# ── Export to JSON ─────────────────────────────────────────────────────────────
@sysadmin_bp.route('/table/<slug>/export-json')
@login_required
@sysadmin_required
def export_json(slug):
    if slug not in TABLE_REGISTRY:
        abort(404)
    meta    = TABLE_REGISTRY[slug]
    model   = meta['model']
    columns = _get_columns(model)
    records = [_record_to_dict(r, columns) for r in model.query.all()]

    fname = f"{slug}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    resp  = make_response(json.dumps(records, indent=2, default=str))
    resp.headers['Content-Type']        = 'application/json'
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    return resp

# ── Delete single record ───────────────────────────────────────────────────────
@sysadmin_bp.route('/table/<slug>/delete/<int:record_id>', methods=['POST'])
@login_required
@sysadmin_required
def delete_record(slug, record_id):
    if slug not in TABLE_REGISTRY:
        abort(404)
    model  = TABLE_REGISTRY[slug]['model']
    record = model.query.get_or_404(record_id)
    db.session.delete(record)
    db.session.commit()
    flash(f'Record #{record_id} deleted from {slug}.', 'success')
    return redirect(url_for('sysadmin.table_view', slug=slug))

# ── Bulk delete ────────────────────────────────────────────────────────────────
@sysadmin_bp.route('/table/<slug>/bulk-delete', methods=['POST'])
@login_required
@sysadmin_required
def bulk_delete(slug):
    if slug not in TABLE_REGISTRY:
        abort(404)
    ids = request.form.getlist('ids', type=int)
    if not ids:
        flash('No records selected.', 'warning')
        return redirect(url_for('sysadmin.table_view', slug=slug))
    model = TABLE_REGISTRY[slug]['model']
    deleted = model.query.filter(model.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()
    flash(f'{deleted} records deleted.', 'success')
    return redirect(url_for('sysadmin.table_view', slug=slug))

# ── Edit record (GET → JSON, POST → save) ─────────────────────────────────────
@sysadmin_bp.route('/table/<slug>/record/<int:record_id>', methods=['GET', 'POST'])
@login_required
@sysadmin_required
def edit_record(slug, record_id):
    if slug not in TABLE_REGISTRY:
        abort(404)
    model   = TABLE_REGISTRY[slug]['model']
    columns = _get_columns(model)
    record  = model.query.get_or_404(record_id)

    if request.method == 'GET':
        return jsonify(_record_to_dict(record, columns))

    # POST: update fields
    for col in columns:
        if col['primary_key']:
            continue
        val = request.form.get(col['name'])
        if val is not None:
            # Type coercion
            if col['type'] in ('Integer', 'BigInteger'):
                try: val = int(val)
                except (ValueError, TypeError): val = None
            elif col['type'] == 'Boolean':
                val = val.lower() in ('true', '1', 'yes', 'on')
            elif val == '':
                val = None
            setattr(record, col['name'], val)
    db.session.commit()
    flash(f'Record #{record_id} updated successfully.', 'success')
    return redirect(url_for('sysadmin.table_view', slug=slug))

# ── Raw SQL Query ──────────────────────────────────────────────────────────────
@sysadmin_bp.route('/sql', methods=['GET', 'POST'])
@login_required
@sysadmin_required
def raw_sql():
    result = None
    columns = []
    error   = None
    query_text = ''

    if request.method == 'POST':
        query_text = request.form.get('query', '').strip()
        try:
            res = db.session.execute(text(query_text))
            if res.returns_rows:
                columns = list(res.keys())
                result  = [list(row) for row in res.fetchall()]
            else:
                db.session.commit()
                flash(f'Query executed successfully. Rows affected: {res.rowcount}', 'success')
        except Exception as e:
            db.session.rollback()
            error = str(e)

    return render_template('sysadmin/sql.html',
        query_text=query_text, result=result,
        columns=columns, error=error,
        table_stats=[{'slug': s, 'label': m['label'], 'icon': m['icon'],
                      'color': m['color'], 'count': '...'}
                     for s, m in TABLE_REGISTRY.items()],
        active_page='sql'
    )

# ── Create System Admin user (one-time setup CLI) ──────────────────────────────
def create_system_admin(app):
    """Call this from Flask CLI: flask create-sysadmin"""
    with app.app_context():
        from app.extensions import bcrypt as _bcrypt
        existing = User.query.filter_by(role=Roles.SYSTEM_ADMIN).first()
        if existing:
            print(f'System admin already exists: {existing.email}')
            return
        pw_hash = _bcrypt.generate_password_hash('SysAdmin@CSMSS2025').decode('utf-8')
        admin = User(
            name='System Administrator',
            email='sysadmin@csmss.edu.in',
            password_hash=pw_hash,
            role=Roles.SYSTEM_ADMIN,
            status=Status.ACTIVE,
            must_change_password=True,
        )
        db.session.add(admin)
        db.session.commit()
        print('[OK] System admin created.')
        print('   Email:    sysadmin@csmss.edu.in')
        print('   Password: SysAdmin@CSMSS2025  (change immediately after login)')
