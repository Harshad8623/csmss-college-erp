from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, abort, send_file
from flask_login import login_required, current_user
from app.extensions import db
from app.models import (
    Attendance, Student, Subject, Class, User, Roles, ApprovalStatus
)
from app.utils.decorators import role_required
from app.utils.helpers import (
    send_notification, calculate_attendance_percentage, classes_needed_for_75,
    get_dept_for_hod, get_class_for_ct, get_tg_student_ids, get_students_for_subject
)
from datetime import date, datetime
import io
import openpyxl

attendance_bp = Blueprint('attendance', __name__, url_prefix='/attendance')


def _scoped_subjects():
    """Return subjects list scoped by current user's role."""
    role = current_user.role
    if role == Roles.SUPER_ADMIN:
        return Subject.query.all()
    elif role == Roles.HOD:
        dept_id = get_dept_for_hod(current_user.id)
        if dept_id:
            class_ids = [c.id for c in Class.query.filter_by(department_id=dept_id).all()]
            return Subject.query.filter(Subject.class_id.in_(class_ids)).all()
        return []
    elif role == Roles.CLASS_TEACHER:
        cls = get_class_for_ct(current_user.id)
        return Subject.query.filter_by(class_id=cls.id).all() if cls else []
    else:
        # Regular teacher — only their assigned subjects
        return Subject.query.filter_by(teacher_id=current_user.id).all()


@attendance_bp.route('/')
@login_required
def index():
    if current_user.role in [Roles.TEACHER, Roles.CLASS_TEACHER, Roles.HOD, Roles.SUPER_ADMIN]:
        subjects = _scoped_subjects()
        subject_ids = [s.id for s in subjects]
        recent_sessions = []
        if subject_ids:
            sessions_query = db.session.query(Attendance.date, Attendance.subject_id).filter(
                Attendance.subject_id.in_(subject_ids)
            ).distinct().order_by(Attendance.date.desc()).limit(30).all()

            for session_date, sub_id in sessions_query:
                sub = Subject.query.get(sub_id)
                recent_sessions.append({'date': session_date, 'subject': sub})

        return render_template('attendance/teacher_view.html',
                               subjects=subjects, recent_sessions=recent_sessions)
    else:
        student = Student.query.filter_by(user_id=current_user.id).first()
        if not student:
            flash('Student profile not found.', 'danger')
            return redirect(url_for('dashboard.index'))
        return _student_attendance(student)


def _student_attendance(student):
    subjects = Subject.query.filter_by(class_id=student.class_id).all()
    attendance_summary = []
    for sub in subjects:
        records = Attendance.query.filter_by(student_id=student.id, subject_id=sub.id)\
                    .order_by(Attendance.date.desc()).all()
        total   = len(records)
        present = sum(1 for r in records if r.status)
        pct     = round((present / total) * 100, 2) if total else 0
        needed  = classes_needed_for_75(present, total)
        attendance_summary.append({
            'subject':      sub,
            'total':        total,
            'present':      present,
            'absent':       total - present,
            'percentage':   pct,
            'is_defaulter': pct < 75,
            'classes_needed': needed,
            'records':      records[:30]
        })
    overall = student.attendance_percentage()
    return render_template('attendance/student_view.html',
        student=student, attendance_summary=attendance_summary, overall=overall)


@attendance_bp.route('/mark', methods=['GET', 'POST'])
@login_required
@role_required(Roles.TEACHER, Roles.CLASS_TEACHER, Roles.HOD, Roles.SUPER_ADMIN)
def mark():
    subjects = _scoped_subjects()

    selected_subject = None
    students = []
    selected_date = date.today()
    existing = {}

    subject_id = request.args.get('subject_id') or request.form.get('subject_id')
    att_date   = request.args.get('date') or request.form.get('date', str(date.today()))

    if subject_id:
        selected_subject = Subject.query.get(subject_id)

        # Verify this subject is in scope
        if selected_subject and selected_subject not in subjects:
            abort(403)

        if selected_subject:
            students = get_students_for_subject(selected_subject)
            selected_date = date.fromisoformat(att_date)
            existing_records = Attendance.query.filter_by(
                subject_id=subject_id, date=selected_date
            ).all()
            existing = {r.student_id: r.status for r in existing_records}

    if request.method == 'POST' and selected_subject:
        selected_date = date.fromisoformat(request.form.get('date', str(date.today())))
        present_ids = [int(x) for x in request.form.getlist('present')]

        # Bulk load existing records
        existing_records = Attendance.query.filter_by(
            subject_id=selected_subject.id, date=selected_date
        ).all()
        existing_dict = {r.student_id: r for r in existing_records}

        for student in students:
            is_present = student.id in present_ids
            if student.id in existing_dict:
                record = existing_dict[student.id]
                record.status = is_present
                record.marked_by = current_user.id
            else:
                db.session.add(Attendance(
                    student_id=student.id,
                    subject_id=selected_subject.id,
                    date=selected_date,
                    status=is_present,
                    marked_by=current_user.id
                ))

        db.session.commit()

        # Bulk notification for defaulters
        student_ids = [s.id for s in students]
        if student_ids:
            stats = db.session.query(
                Attendance.student_id,
                db.func.count(Attendance.id).label('total'),
                db.func.sum(db.case((Attendance.status == True, 1), else_=0)).label('present')
            ).filter(
                Attendance.subject_id == selected_subject.id,
                Attendance.student_id.in_(student_ids)
            ).group_by(Attendance.student_id).all()
            
            stats_dict = {r.student_id: {'total': r.total, 'present': r.present or 0} for r in stats}
            
            from app.models import Notification
            notifs = []
            for student in students:
                st = stats_dict.get(student.id)
                if st and st['total'] > 0:
                    pct = round((st['present'] / st['total']) * 100, 2)
                    if pct < 75:
                        notifs.append(Notification(
                            user_id=student.user_id,
                            message=f'⚠️ Your attendance in {selected_subject.name} is {pct}% — below 75%!',
                            type='warning',
                            link=url_for('attendance.index')
                        ))
            
            if notifs:
                db.session.bulk_save_objects(notifs)
                db.session.commit()

        flash(f'Attendance saved — {selected_date.strftime("%d %b %Y")} · {selected_subject.name}', 'success')
        return redirect(url_for('attendance.mark',
            subject_id=selected_subject.id, date=str(selected_date)))

    return render_template('attendance/mark.html',
        subjects=subjects, selected_subject=selected_subject,
        students=students, selected_date=selected_date,
        existing=existing, att_date=att_date)


@attendance_bp.route('/report/<int:class_id>')
@login_required
@role_required(Roles.TEACHER, Roles.CLASS_TEACHER, Roles.HOD, Roles.SUPER_ADMIN, Roles.CR)
def class_report(class_id):
    class_ = Class.query.get_or_404(class_id)

    # Scope check
    if current_user.role == Roles.HOD:
        dept_id = get_dept_for_hod(current_user.id)
        if class_.department_id != dept_id:
            abort(403)
    elif current_user.role == Roles.CLASS_TEACHER:
        cls = get_class_for_ct(current_user.id)
        if not cls or cls.id != class_id:
            abort(403)
    elif current_user.role == Roles.CR:
        student = Student.query.filter_by(user_id=current_user.id).first()
        if not student or student.class_id != class_id:
            abort(403)
    elif current_user.role == Roles.TEACHER:
        taught_class_ids = [s.class_id for s in Subject.query.filter_by(teacher_id=current_user.id).all() if s.class_id]
        tg_class_ids = [s.class_id for s in Student.query.filter_by(tg_id=current_user.id).all() if s.class_id]
        if class_id not in taught_class_ids and class_id not in tg_class_ids:
            abort(403)

    students = Student.query.filter_by(
        class_id=class_id, 
        approval_status=ApprovalStatus.APPROVED
    ).order_by(Student.roll_no.asc()).all()
    subjects = Subject.query.filter_by(class_id=class_id).all()

    student_ids = [s.id for s in students]
    subject_ids = [s.id for s in subjects]
    
    import collections
    stats_dict = collections.defaultdict(dict)
    if student_ids and subject_ids:
        stats = db.session.query(
            Attendance.student_id,
            Attendance.subject_id,
            db.func.count(Attendance.id).label('total'),
            db.func.sum(db.case((Attendance.status == True, 1), else_=0)).label('present')
        ).filter(
            Attendance.student_id.in_(student_ids),
            Attendance.subject_id.in_(subject_ids)
        ).group_by(Attendance.student_id, Attendance.subject_id).all()
        
        for r in stats:
            stats_dict[r.student_id][r.subject_id] = {'total': r.total, 'present': r.present or 0}

    report = []
    for student in students:
        row = {'student': student, 'subjects': {}}
        overall_total = overall_present = 0
        for sub in subjects:
            if sub.is_elective and student not in sub.enrolled_students:
                row['subjects'][sub.id] = {'total': '-', 'present': '-', 'pct': '-'}
                continue

            st = stats_dict[student.id].get(sub.id, {'total': 0, 'present': 0})
            total = st['total']
            present = st['present']
            pct = round((present / total) * 100, 2) if total else 0
            row['subjects'][sub.id] = {'total': total, 'present': present, 'pct': pct}
            overall_total   += total
            overall_present += present
        row['overall_pct']  = round((overall_present / overall_total) * 100, 2) if overall_total else 0
        row['is_defaulter'] = row['overall_pct'] < 75
        report.append(row)

    return render_template('attendance/class_report.html',
        class_=class_, students=students, subjects=subjects, report=report)


@attendance_bp.route('/api/chart/<int:student_id>')
@login_required
def chart_data(student_id):
    student  = Student.query.get_or_404(student_id)

    # Authorization: students can only see their own chart
    if current_user.role in [Roles.STUDENT, Roles.CR]:
        my_student = Student.query.filter_by(user_id=current_user.id).first()
        if not my_student or my_student.id != student_id:
            abort(403)
    # Staff scope check
    elif current_user.role == Roles.TEACHER:
        taught_class_ids = [s.class_id for s in Subject.query.filter_by(teacher_id=current_user.id).all() if s.class_id]
        is_tg = (student.tg_id == current_user.id)
        if student.class_id not in taught_class_ids and not is_tg:
            abort(403)
    elif current_user.role == Roles.CLASS_TEACHER:
        cls = get_class_for_ct(current_user.id)
        if not cls or student.class_id != cls.id:
            abort(403)
    elif current_user.role == Roles.HOD:
        dept_id = get_dept_for_hod(current_user.id)
        cls = Class.query.get(student.class_id)
        if not cls or cls.department_id != dept_id:
            abort(403)

    subjects = Subject.query.filter_by(class_id=student.class_id).all()
    labels, data = [], []
    for sub in subjects:
        pct = calculate_attendance_percentage(student_id, sub.id)
        labels.append(sub.name)
        data.append(pct)
    return jsonify({'labels': labels, 'data': data})


# ── Class Teacher: Full Date-wise Attendance View ────────────────────────────
@attendance_bp.route('/ct-view')
@login_required
@role_required(Roles.CLASS_TEACHER, Roles.HOD, Roles.SUPER_ADMIN)
def ct_view():
    """Date-wise attendance view across all subjects for the class teacher's class."""
    from app.utils.helpers import get_class_for_ct, get_dept_for_hod

    # Resolve which class to show
    if current_user.role == Roles.CLASS_TEACHER:
        cls = get_class_for_ct(current_user.id)
        if not cls:
            flash('You are not assigned as a Class Teacher for any class.', 'warning')
            return redirect(url_for('attendance.index'))
    elif current_user.role == Roles.HOD:
        # HOD picks via query param, defaults to first class in dept
        dept_id = get_dept_for_hod(current_user.id)
        class_id_param = request.args.get('class_id', type=int)
        if class_id_param:
            cls = Class.query.get_or_404(class_id_param)
            if cls.department_id != dept_id:
                abort(403)
        else:
            cls = Class.query.filter_by(department_id=dept_id).first()
            if not cls:
                flash('No classes found in your department.', 'warning')
                return redirect(url_for('attendance.index'))
    else:  # SUPER_ADMIN
        class_id_param = request.args.get('class_id', type=int)
        if class_id_param:
            cls = Class.query.get_or_404(class_id_param)
        else:
            cls = Class.query.first()
            if not cls:
                flash('No classes found.', 'warning')
                return redirect(url_for('attendance.index'))

    subjects = Subject.query.filter_by(class_id=cls.id).order_by(Subject.name).all()
    students_count = Student.query.filter_by(
        class_id=cls.id, approval_status=ApprovalStatus.APPROVED
    ).count()

    # Build per-subject date-wise session data
    subject_data = []
    for sub in subjects:
        # All distinct dates this subject has attendance records
        sessions_raw = db.session.query(
            Attendance.date,
            db.func.count(Attendance.id).label('total'),
            db.func.sum(db.case((Attendance.status == True, 1), else_=0)).label('present')
        ).filter(
            Attendance.subject_id == sub.id
        ).group_by(Attendance.date).order_by(Attendance.date.desc()).all()

        sessions = []
        for row in sessions_raw:
            absent = row.total - (row.present or 0)
            pct = round(((row.present or 0) / row.total) * 100) if row.total else 0
            sessions.append({
                'date': row.date,
                'total': row.total,
                'present': row.present or 0,
                'absent': absent,
                'pct': pct,
            })

        # Overall stats for this subject
        total_lectures = len(sessions)
        if sessions:
            all_total   = sum(s['total'] for s in sessions)
            all_present = sum(s['present'] for s in sessions)
            avg_pct = round((all_present / all_total) * 100) if all_total else 0
        else:
            avg_pct = 0

        subject_data.append({
            'subject': sub,
            'sessions': sessions,
            'total_lectures': total_lectures,
            'avg_pct': avg_pct,
            'teacher': User.query.get(sub.teacher_id) if sub.teacher_id else None,
        })

    # All classes (for HOD/SUPER_ADMIN switcher)
    if current_user.role == Roles.HOD:
        dept_id = get_dept_for_hod(current_user.id)
        all_classes = Class.query.filter_by(department_id=dept_id).all()
    elif current_user.role == Roles.SUPER_ADMIN:
        all_classes = Class.query.all()
    else:
        all_classes = [cls]

    return render_template('attendance/ct_attendance.html',
        cls=cls,
        subject_data=subject_data,
        students_count=students_count,
        all_classes=all_classes,
    )

@attendance_bp.route('/export-template/<int:subject_id>')
@login_required
@role_required(Roles.TEACHER, Roles.CLASS_TEACHER, Roles.HOD, Roles.SUPER_ADMIN)
def export_template(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    # Scope check
    subjects = _scoped_subjects()
    if subject not in subjects:
        abort(403)

    students = get_students_for_subject(subject)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Template"
    
    # Headers
    headers = ['Roll No', 'PRN', 'Student Name', date.today().strftime('%Y-%m-%d')]
    ws.append(headers)
    
    # Format Headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Column Widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    
    # Data Rows
    for student in students:
        ws.append([
            student.roll_no or '',
            student.prn or '',
            student.user.name if student.user else 'Unknown',
            '' # Empty cell for attendance
        ])
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"Attendance_Template_{subject.name[:15]}_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(out, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@attendance_bp.route('/upload-excel', methods=['POST'])
@login_required
@role_required(Roles.TEACHER, Roles.CLASS_TEACHER, Roles.HOD, Roles.SUPER_ADMIN)
def upload_excel():
    subject_id = request.form.get('subject_id')
    file = request.files.get('excel_file')
    
    if not subject_id or not file or not file.filename.endswith('.xlsx'):
        flash('Invalid subject or file type. Please upload a .xlsx file.', 'danger')
        return redirect(url_for('attendance.mark', subject_id=subject_id))
        
    subject = Subject.query.get_or_404(subject_id)
    if subject not in _scoped_subjects():
        abort(403)
        
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        # Read headers
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            flash('Empty template or missing student data.', 'danger')
            return redirect(url_for('attendance.mark', subject_id=subject_id))
            
        headers = rows[0]
        # Find date columns (typically anything from column index 3 onwards)
        # Expected headers: Roll No, PRN, Student Name, <Date 1>, <Date 2>...
        date_cols = {}
        for idx, h in enumerate(headers):
            if idx >= 3 and h: # Column 4 onwards
                try:
                    # Attempt to parse as date (could be datetime object if openpyxl parsed it)
                    if isinstance(h, datetime):
                        dt = h.date()
                    else:
                        dt = date.fromisoformat(str(h).strip().split(' ')[0])
                    date_cols[idx] = dt
                except (ValueError, TypeError, AttributeError):
                    flash(f"Invalid date format in header '{h}'. Use YYYY-MM-DD.", "warning")
                    return redirect(url_for('attendance.mark', subject_id=subject_id))
        
        if not date_cols:
            flash('No valid date columns found. Add dates like YYYY-MM-DD as headers.', 'danger')
            return redirect(url_for('attendance.mark', subject_id=subject_id))
            
        students = get_students_for_subject(subject)
        student_map = {}
        for s in students:
            if s.prn: student_map[str(s.prn).strip()] = s
            if s.roll_no: student_map[str(s.roll_no).strip()] = s
            
        # Parse data
        records_to_add = []
        updates_count = 0
        new_count = 0
        
        for row in rows[1:]:
            roll_no = str(row[0]).strip() if row[0] else ''
            prn = str(row[1]).strip() if row[1] else ''
            
            student = student_map.get(prn) or student_map.get(roll_no)
            if not student:
                continue # Skip if student not found in this class/subject
                
            for col_idx, dt in date_cols.items():
                val = str(row[col_idx]).strip().upper() if len(row) > col_idx and row[col_idx] is not None else ''
                if val not in ['P', 'A', 'PRESENT', 'ABSENT']:
                    continue
                    
                is_present = val in ['P', 'PRESENT']
                
                # Check existing record
                existing = Attendance.query.filter_by(
                    student_id=student.id, subject_id=subject.id, date=dt
                ).first()
                
                if existing:
                    if existing.status != is_present:
                        existing.status = is_present
                        existing.marked_by = current_user.id
                        updates_count += 1
                else:
                    db.session.add(Attendance(
                        student_id=student.id,
                        subject_id=subject.id,
                        date=dt,
                        status=is_present,
                        marked_by=current_user.id
                    ))
                    new_count += 1
                    
        db.session.commit()
        flash(f'Excel upload complete! Added {new_count} records and updated {updates_count} records.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error processing Excel file: {str(e)}', 'danger')
        
    return redirect(url_for('attendance.mark', subject_id=subject_id))

