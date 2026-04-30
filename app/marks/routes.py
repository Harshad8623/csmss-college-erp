from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, abort
from flask_login import login_required, current_user
from app.extensions import db
from app.models import Marks, Student, Subject, Class, User, Roles, ApprovalStatus, ExamType
from app.utils.decorators import role_required
from app.utils.helpers import send_notification, get_grade, get_dept_for_hod, get_class_for_ct, get_students_for_subject
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import send_file

marks_bp = Blueprint('marks', __name__, url_prefix='/marks')

def _scoped_subjects():
    """Return subjects list scoped by current user's role."""
    role = current_user.role
    if role == Roles.SUPER_ADMIN:
        return Subject.query.all()
    elif role == Roles.HOD:
        dept_id = get_dept_for_hod(current_user.id)
        if dept_id:
            class_ids = [c.id for c in Class.query.filter_by(department_id=dept_id).all()]
            return Subject.query.filter(Subject.class_id.in_(class_ids)).all()
        return []
    elif role == Roles.CLASS_TEACHER:
        cls = get_class_for_ct(current_user.id)
        return Subject.query.filter_by(class_id=cls.id).all() if cls else []
    else:
        return Subject.query.filter_by(teacher_id=current_user.id).all()

@marks_bp.route('/')
@login_required
def index():
    if current_user.role in [Roles.TEACHER, Roles.CLASS_TEACHER, Roles.HOD, Roles.SUPER_ADMIN]:
        subjects = _scoped_subjects()
        return render_template('marks/teacher_view.html', subjects=subjects)
    else:
        student = Student.query.filter_by(user_id=current_user.id).first()
        if not student:
            flash('Student profile not found.', 'danger')
            return redirect(url_for('dashboard.index'))
        all_marks = Marks.query.filter_by(student_id=student.id)\
            .order_by(Marks.created_at.desc()).all()
        # Group by subject
        subjects_summary = {}
        for m in all_marks:
            sid = m.subject_id
            if sid not in subjects_summary:
                subjects_summary[sid] = {'subject': m.subject, 'marks': []}
            subjects_summary[sid]['marks'].append(m)
        return render_template('marks/student_view.html',
            student=student, subjects_summary=subjects_summary.values())

@marks_bp.route('/upload', methods=['GET', 'POST'])
@login_required
@role_required(Roles.TEACHER, Roles.CLASS_TEACHER, Roles.HOD, Roles.SUPER_ADMIN)
def upload():
    subjects = _scoped_subjects()

    selected_subject = None
    students = []
    exam_type = request.args.get('exam_type', ExamType.CT1)
    subject_id = request.args.get('subject_id') or request.form.get('subject_id')

    if subject_id:
        selected_subject = Subject.query.get(subject_id)
        # Scope check — block access to out-of-scope subjects
        if selected_subject and selected_subject not in subjects:
            abort(403)
        exam_type = request.args.get('exam_type', ExamType.CT1) or request.form.get('exam_type', ExamType.CT1)
        if selected_subject:
            students = get_students_for_subject(selected_subject)

    if request.method == 'POST' and selected_subject:
        exam_type = request.form.get('exam_type', ExamType.CT1)
        max_marks = float(request.form.get('max_marks', 100))
        count = 0
        for student in students:
            marks_val = request.form.get(f'marks_{student.id}', '').strip()
            if marks_val == '':
                continue
            try:
                marks_float = float(marks_val)
            except ValueError:
                continue
            # Upsert
            existing = Marks.query.filter_by(
                student_id=student.id,
                subject_id=selected_subject.id,
                exam_type=exam_type
            ).first()
            if existing:
                existing.marks = marks_float
                existing.max_marks = max_marks
                existing.uploaded_by = current_user.id
            else:
                m = Marks(student_id=student.id, subject_id=selected_subject.id,
                          marks=marks_float, max_marks=max_marks, exam_type=exam_type,
                          uploaded_by=current_user.id)
                db.session.add(m)
            count += 1
            send_notification(
                student.user_id,
                f'📊 Marks uploaded for {selected_subject.name} — {exam_type.title()}',
                'info', url_for('marks.index')
            )

        db.session.commit()
        flash(f'Marks uploaded for {count} students — {selected_subject.name} [{exam_type}]', 'success')
        return redirect(url_for('marks.upload', subject_id=selected_subject.id, exam_type=exam_type))

    # Load existing marks for display
    existing_marks = {}
    if selected_subject:
        for m in Marks.query.filter_by(subject_id=selected_subject.id, exam_type=exam_type).all():
            existing_marks[m.student_id] = m

    return render_template('marks/upload.html',
        subjects=subjects, selected_subject=selected_subject,
        students=students, exam_type=exam_type,
        existing_marks=existing_marks,
        exam_types=ExamType.ALL,
        exam_labels=ExamType.LABELS,
    )

@marks_bp.route('/export-template/<int:subject_id>/<exam_type>')
@login_required
@role_required(Roles.TEACHER, Roles.CLASS_TEACHER, Roles.HOD, Roles.SUPER_ADMIN)
def export_template(subject_id, exam_type):
    subject = Subject.query.get_or_404(subject_id)
    subjects_scope = _scoped_subjects()
    if subject not in subjects_scope:
        abort(403)
    
    students = get_students_for_subject(subject)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Marks_{exam_type}"
    
    headers = ['Roll No', 'PRN', 'Student Name', 'Max Marks', 'Marks Obtained']
    ws.append(headers)
    
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    for student in students:
        ws.append([
            student.roll_no or '',
            student.prn or '',
            student.user.name if student.user else 'Unknown',
            100,  # Default Max Marks
            ''    # Marks Obtained to be filled by teacher
        ])
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"Marks_Template_{subject.code or subject.id}_{exam_type}.xlsx"
    return send_file(out, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@marks_bp.route('/upload-excel', methods=['POST'])
@login_required
@role_required(Roles.TEACHER, Roles.CLASS_TEACHER, Roles.HOD, Roles.SUPER_ADMIN)
def upload_excel():
    subject_id = request.form.get('subject_id')
    exam_type = request.form.get('exam_type')
    
    if not subject_id or not exam_type:
        flash('Missing subject or exam type.', 'danger')
        return redirect(url_for('marks.upload'))
        
    subject = Subject.query.get_or_404(subject_id)
    if subject not in _scoped_subjects():
        abort(403)
        
    if 'file' not in request.files:
        flash('No file part', 'danger')
        return redirect(url_for('marks.upload', subject_id=subject_id, exam_type=exam_type))
        
    file = request.files['file']
    if file.filename == '':
        flash('No selected file', 'danger')
        return redirect(url_for('marks.upload', subject_id=subject_id, exam_type=exam_type))
        
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
        
        prn_idx = -1
        roll_idx = -1
        max_idx = -1
        marks_idx = -1
        
        for i, h in enumerate(headers):
            if 'prn' in h: prn_idx = i
            elif 'roll' in h: roll_idx = i
            elif 'max' in h: max_idx = i
            elif 'obtain' in h or 'marks' in h and not 'max' in h: marks_idx = i
            
        if (prn_idx == -1 and roll_idx == -1) or marks_idx == -1:
            flash('Invalid template. Missing PRN/Roll No or Marks column.', 'danger')
            return redirect(url_for('marks.upload', subject_id=subject_id, exam_type=exam_type))
            
        students = {s.prn: s for s in get_students_for_subject(subject) if s.prn}
        roll_students = {s.roll_no: s for s in get_students_for_subject(subject) if s.roll_no}
        
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            prn = str(row[prn_idx]).strip() if prn_idx != -1 and row[prn_idx] is not None else None
            roll = str(row[roll_idx]).strip() if roll_idx != -1 and row[roll_idx] is not None else None
            
            marks_val = row[marks_idx]
            max_val = row[max_idx] if max_idx != -1 and row[max_idx] is not None else 100
            
            if marks_val is None or str(marks_val).strip() == '':
                continue
                
            try:
                marks_float = float(marks_val)
                max_float = float(max_val)
            except ValueError:
                continue
                
            student = students.get(prn) or roll_students.get(roll)
            if not student:
                continue
                
            existing = Marks.query.filter_by(
                student_id=student.id,
                subject_id=subject.id,
                exam_type=exam_type
            ).first()
            
            if existing:
                existing.marks = marks_float
                existing.max_marks = max_float
                existing.uploaded_by = current_user.id
            else:
                m = Marks(student_id=student.id, subject_id=subject.id,
                          marks=marks_float, max_marks=max_float, exam_type=exam_type,
                          uploaded_by=current_user.id)
                db.session.add(m)
            count += 1
            send_notification(
                student.user_id,
                f'📊 Marks uploaded for {subject.name} — {exam_type.title()}',
                'info', url_for('marks.index')
            )
            
        db.session.commit()
        flash(f'Successfully imported marks for {count} students.', 'success')
        
    except Exception as e:
        flash(f'Error processing file: {str(e)}', 'danger')
        
    return redirect(url_for('marks.upload', subject_id=subject_id, exam_type=exam_type))

@marks_bp.route('/report/<int:class_id>')
@login_required
@role_required(Roles.TEACHER, Roles.CLASS_TEACHER, Roles.HOD, Roles.SUPER_ADMIN)
def class_report(class_id):
    class_ = Class.query.get_or_404(class_id)
    # Scope check
    if current_user.role == Roles.HOD:
        dept_id = get_dept_for_hod(current_user.id)
        if class_.department_id != dept_id:
            abort(403)
    elif current_user.role == Roles.CLASS_TEACHER:
        cls = get_class_for_ct(current_user.id)
        if not cls or cls.id != class_id:
            abort(403)
    students = Student.query.filter_by(class_id=class_id, approval_status=ApprovalStatus.APPROVED).all()
    subjects = Subject.query.filter_by(class_id=class_id).all()
    exam_type = request.args.get('exam_type', ExamType.CT1)

    report = []
    student_ids = [s.id for s in students]
    subject_ids = [s.id for s in subjects]
    
    import collections
    marks_dict = collections.defaultdict(dict)
    
    if student_ids and subject_ids:
        all_marks = Marks.query.filter(
            Marks.student_id.in_(student_ids),
            Marks.subject_id.in_(subject_ids),
            Marks.exam_type == exam_type
        ).all()
        for m in all_marks:
            marks_dict[m.student_id][m.subject_id] = m

    for student in students:
        row = {'student': student, 'subjects': {}, 'total': 0, 'max_total': 0}
        for sub in subjects:
            if sub.is_elective and student not in sub.enrolled_students:
                row['subjects'][sub.id] = 'N/A'
                continue
            
            m = marks_dict[student.id].get(sub.id)
            row['subjects'][sub.id] = m
            if m:
                row['total'] += m.marks
                row['max_total'] += m.max_marks
        row['percentage'] = round((row['total'] / row['max_total']) * 100, 2) if row['max_total'] else 0
        row['grade_info'] = get_grade(row['percentage'])
        report.append(row)

    # Sort by percentage descending for rank
    report.sort(key=lambda x: x['percentage'], reverse=True)
    for i, r in enumerate(report):
        r['rank'] = i + 1

    return render_template('marks/class_report.html',
        class_=class_, subjects=subjects, report=report, exam_type=exam_type,
        exam_types=ExamType.ALL,
        exam_labels=ExamType.LABELS,
    )

@marks_bp.route('/api/performance/<int:student_id>')
@login_required
def performance_data(student_id):
    student = Student.query.get_or_404(student_id)
    all_class_subjects = Subject.query.filter_by(class_id=student.class_id).all()
    subjects = [s for s in all_class_subjects if not s.is_elective or student in s.enrolled_students]
    
    labels = [s.name for s in subjects]
    datasets = {}
    for et in ExamType.ALL:
        datasets[et] = []
        for sub in subjects:
            m = Marks.query.filter_by(
                student_id=student_id, subject_id=sub.id, exam_type=et
            ).first()
            datasets[et].append(m.percentage if m else 0)
    return jsonify({'labels': labels, 'datasets': datasets, 'exam_labels': ExamType.LABELS})
