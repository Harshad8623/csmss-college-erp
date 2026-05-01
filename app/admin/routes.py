from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, send_file, current_app
from flask_login import login_required, current_user
from app.extensions import db, bcrypt
from app.models import (
    User, Student, Teacher, Department, Class, Subject,
    Roles, Status, ApprovalStatus
)
from app.utils.decorators import role_required
from app.utils.helpers import get_dept_for_hod, get_class_for_ct
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

# ── Helpers ──────────────────────────────────────────────────────────────────
def _hod_dept_id():
    """Return department_id for the current HOD, or None."""
    return get_dept_for_hod(current_user.id)

def _ct_class():
    """Return Class object for the current Class Teacher, or None."""
    return get_class_for_ct(current_user.id)

# ── User Management ─────────────────────────────────────────────────────────
@admin_bp.route('/users')
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD)
def users():
    filter_role   = request.args.get('role', '')
    filter_status = request.args.get('status', '')
    query = User.query

    # HOD sees only teachers/staff in their own department
    if current_user.role == Roles.HOD:
        dept_id = _hod_dept_id()
        if dept_id:
            teacher_user_ids = [
                t.user_id for t in Teacher.query.filter_by(department_id=dept_id).all()
            ]
            query = query.filter(User.id.in_(teacher_user_ids))
        else:
            query = query.filter(User.id == 0)  # No dept → empty

    if filter_role:
        query = query.filter_by(role=filter_role)
    if filter_status:
        query = query.filter_by(status=filter_status)

    users = query.order_by(User.created_at.desc()).all()
    departments = Department.query.all()
    return render_template('admin/users.html', users=users, roles=Roles.ALL,
                           filter_role=filter_role, filter_status=filter_status,
                           departments=departments)


@admin_bp.route('/users/<int:id>/activate', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER)
def activate_user(id):
    user = User.query.get_or_404(id)
    # HOD can only activate users from own department
    if current_user.role == Roles.HOD:
        dept_id = _hod_dept_id()
        teacher = Teacher.query.filter_by(user_id=user.id).first()
        if not teacher or teacher.department_id != dept_id:
            abort(403)
    # Class Teacher can only activate students from their class
    elif current_user.role == Roles.CLASS_TEACHER:
        cls = _ct_class()
        if not cls:
            abort(403)
        student = Student.query.filter_by(user_id=user.id).first()
        if not student or student.class_id != cls.id:
            abort(403)
    user.status = Status.ACTIVE
    db.session.commit()
    flash(f'{user.name} activated.', 'success')
    return redirect(request.referrer or url_for('admin.users'))


@admin_bp.route('/users/<int:id>/block', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN)
def block_user(id):
    user = User.query.get_or_404(id)
    user.status = Status.BLOCKED
    db.session.commit()
    flash(f'{user.name} blocked.', 'warning')
    return redirect(url_for('admin.users'))


@admin_bp.route('/users/<int:id>/change-role', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN)
def change_role(id):
    user = User.query.get_or_404(id)
    # Prevent changing your own role (could lock yourself out)
    if user.id == current_user.id:
        flash('You cannot change your own role.', 'danger')
        return redirect(url_for('admin.users'))
    new_role = request.form.get('role')
    if new_role in Roles.ALL:
        user.role = new_role
        db.session.commit()
        flash(f'{user.name} role changed to {new_role}.', 'success')
    return redirect(url_for('admin.users'))


@admin_bp.route('/users/<int:id>/unblock', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN)
def unblock_user(id):
    user = User.query.get_or_404(id)
    user.status = Status.ACTIVE
    db.session.commit()
    flash(f'{user.name} has been unblocked and is now active.', 'success')
    return redirect(request.referrer or url_for('admin.users'))


@admin_bp.route('/users/<int:id>/delete', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN)
def delete_user(id):
    from app.models import (
        Notification, Message, AuditLog, Attendance, AbsenteeReason,
        Marks, Grievance, Certificate, Assignment, AssignmentSubmission,
        LeaveApplication, Notice, PracticalSession, EventSession,
        PracticalRecord, EventRecord
    )
    user = User.query.get_or_404(id)

    # Prevent self-deletion
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('admin.users'))

    name = user.name

    try:
        # ── 1. Nullify FK references where column is nullable ──────────────
        # Departments where this user is HOD
        Department.query.filter_by(hod_id=user.id).update({'hod_id': None})
        # Classes where this user is Class Teacher
        Class.query.filter_by(class_teacher_id=user.id).update({'class_teacher_id': None})
        # Subjects where this user is teacher
        Subject.query.filter_by(teacher_id=user.id).update({'teacher_id': None})
        # Students whose TG is this user
        Student.query.filter_by(tg_id=user.id).update({'tg_id': None})
        # Students approved by this user
        Student.query.filter_by(approved_by=user.id).update({'approved_by': None})
        # Grievances assigned to this user
        Grievance.query.filter_by(assigned_to=user.id).update({'assigned_to': None})
        # Certificates approved by this user
        Certificate.query.filter_by(approved_by=user.id).update({'approved_by': None})
        # Leave applications approved by this user (TG / CT)
        LeaveApplication.query.filter_by(tg_approved_by=user.id).update({'tg_approved_by': None})
        LeaveApplication.query.filter_by(ct_approved_by=user.id).update({'ct_approved_by': None})
        # Notices posted by this user
        Notice.query.filter_by(posted_by=user.id).update({'posted_by': None})
        # Assignments created by this user
        Assignment.query.filter_by(created_by=user.id).update({'created_by': None})
        # Attendance marked_by this user
        Attendance.query.filter_by(marked_by=user.id).update({'marked_by': None})
        # Marks uploaded_by this user
        Marks.query.filter_by(uploaded_by=user.id).update({'uploaded_by': None})
        # PracticalSession / EventSession marked_by this user
        PracticalSession.query.filter_by(marked_by=user.id).update({'marked_by': None})
        EventSession.query.filter_by(marked_by=user.id).update({'marked_by': None})
        # AbsenteeReason requested_by this user
        AbsenteeReason.query.filter_by(requested_by=user.id).update({'requested_by': None})
        # AuditLog entries by this user
        AuditLog.query.filter_by(user_id=user.id).update({'user_id': None})

        db.session.flush()

        # ── 2. Delete owned rows (non-nullable FK) ─────────────────────────
        # Notifications
        Notification.query.filter_by(user_id=user.id).delete()
        # Messages sent/received
        Message.query.filter(
            db.or_(Message.sender_id == user.id, Message.receiver_id == user.id)
        ).delete(synchronize_session=False)

        db.session.flush()

        # ── 3. Delete student profile and all its children first ───────────
        if user.student_profile:
            student = user.student_profile
            # Delete student's AbsenteeReasons FIRST (FK → attendance.id)
            from app.models import AbsenteeReason
            att_ids = [a.id for a in Attendance.query.filter_by(student_id=student.id).all()]
            if att_ids:
                AbsenteeReason.query.filter(AbsenteeReason.attendance_id.in_(att_ids)).delete(synchronize_session=False)
            Attendance.query.filter_by(student_id=student.id).delete()
            Marks.query.filter_by(student_id=student.id).delete()
            AssignmentSubmission.query.filter_by(student_id=student.id).delete()
            PracticalRecord.query.filter_by(student_id=student.id).delete()
            EventRecord.query.filter_by(student_id=student.id).delete()
            # Delete GrievanceReplies FIRST (FK → grievances.id)
            from app.models import GrievanceReply
            griev_ids = [g.id for g in Grievance.query.filter_by(student_id=student.id).all()]
            if griev_ids:
                GrievanceReply.query.filter(GrievanceReply.grievance_id.in_(griev_ids)).delete(synchronize_session=False)
            Grievance.query.filter_by(student_id=student.id).delete()
            Certificate.query.filter_by(student_id=student.id).delete()
            LeaveApplication.query.filter_by(student_id=student.id).delete()
            db.session.flush()
            db.session.delete(student)
            db.session.flush()

        # ── 4. Delete teacher profile ──────────────────────────────────────
        if user.teacher_profile:
            db.session.delete(user.teacher_profile)
            db.session.flush()

        # ── 5. Finally delete the user ─────────────────────────────────────
        db.session.delete(user)
        db.session.commit()
        flash(f'User "{name}" has been permanently deleted.', 'danger')

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error deleting user {id}: {e}')
        flash(f'Could not delete user: {e}', 'danger')

    return redirect(url_for('admin.users'))


# ── Student Approvals ────────────────────────────────────────────────────────
@admin_bp.route('/students/approve/<int:id>', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER)
def approve_student(id):
    student = Student.query.get_or_404(id)

    # Scope check
    if current_user.role == Roles.HOD:
        dept_id = _hod_dept_id()
        cls = Class.query.get(student.class_id)
        if not cls or cls.department_id != dept_id:
            abort(403)
    elif current_user.role == Roles.CLASS_TEACHER:
        cls = _ct_class()
        if not cls or student.class_id != cls.id:
            abort(403)

    action = request.form.get('action', 'approve')
    student.approval_status = ApprovalStatus.APPROVED if action == 'approve' else ApprovalStatus.REJECTED
    student.approved_by = current_user.id
    student.user.status = Status.ACTIVE if action == 'approve' else Status.BLOCKED
    db.session.commit()
    flash(f'Student {student.user.name} {student.approval_status}.', 'success')
    from app.utils.helpers import send_notification
    msg = '✅ Your account has been approved! You can now login.' if action == 'approve' else '❌ Your registration was rejected.'
    send_notification(student.user_id, msg, 'success' if action == 'approve' else 'danger')
    return redirect(request.referrer or url_for('dashboard.index'))


# ── Assign TG to Students (Bulk) ─────────────────────────────────────────────
@admin_bp.route('/students/bulk-assign-tg', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER)
def bulk_assign_tg():
    student_ids_str = request.form.get('student_ids', '')
    tg_id = request.form.get('tg_id') or None
    if not student_ids_str:
        flash('No students selected.', 'warning')
        return redirect(request.referrer or url_for('admin.students'))

    student_ids = [int(s.strip()) for s in student_ids_str.split(',') if s.strip().isdigit()]
    if not student_ids:
        flash('Invalid students selected.', 'warning')
        return redirect(request.referrer or url_for('admin.students'))

    students = Student.query.filter(Student.id.in_(student_ids)).all()
    count = 0
    for student in students:
        # Scope check per student
        if current_user.role == Roles.HOD:
            dept_id = _hod_dept_id()
            cls = Class.query.get(student.class_id)
            if not cls or cls.department_id != dept_id:
                continue # Skip out-of-scope students
        elif current_user.role == Roles.CLASS_TEACHER:
            cls = _ct_class()
            if not cls or student.class_id != cls.id:
                continue

        student.tg_id = int(tg_id) if tg_id else None
        count += 1

    db.session.commit()
    flash(f'Teacher Guardian updated for {count} student(s).', 'success')
    return redirect(request.referrer or url_for('admin.students'))


# ── Bulk Delete Students ──────────────────────────────────────────────────────
@admin_bp.route('/students/bulk-delete', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER)
def bulk_delete_students():
    student_ids_str = request.form.get('student_ids', '')
    if not student_ids_str:
        flash('No students selected.', 'warning')
        return redirect(request.referrer or url_for('admin.students'))

    student_ids = [int(s.strip()) for s in student_ids_str.split(',') if s.strip().isdigit()]
    if not student_ids:
        flash('Invalid selection.', 'warning')
        return redirect(request.referrer or url_for('admin.students'))

    students = Student.query.filter(Student.id.in_(student_ids)).all()
    deleted = 0
    for student in students:
        # Scope check
        if current_user.role == Roles.HOD:
            dept_id = _hod_dept_id()
            cls = Class.query.get(student.class_id)
            if not cls or cls.department_id != dept_id:
                continue
        elif current_user.role == Roles.CLASS_TEACHER:
            cls = _ct_class()
            if not cls or student.class_id != cls.id:
                continue

        # Must delete child records FIRST to avoid FK constraint violations in PostgreSQL
        from app.models import (
            Attendance, Marks, AssignmentSubmission,
            PracticalRecord, EventRecord, Grievance,
            Certificate, LeaveApplication
        )
        # Must delete AbsenteeReasons FIRST (FK → attendance.id)
        from app.models import AbsenteeReason
        att_ids = [a.id for a in Attendance.query.filter_by(student_id=student.id).all()]
        if att_ids:
            AbsenteeReason.query.filter(AbsenteeReason.attendance_id.in_(att_ids)).delete(synchronize_session=False)
        Attendance.query.filter_by(student_id=student.id).delete()
        Marks.query.filter_by(student_id=student.id).delete()
        AssignmentSubmission.query.filter_by(student_id=student.id).delete()
        PracticalRecord.query.filter_by(student_id=student.id).delete()
        EventRecord.query.filter_by(student_id=student.id).delete()
        # Delete GrievanceReplies FIRST (FK → grievances.id)
        from app.models import GrievanceReply
        griev_ids = [g.id for g in Grievance.query.filter_by(student_id=student.id).all()]
        if griev_ids:
            GrievanceReply.query.filter(GrievanceReply.grievance_id.in_(griev_ids)).delete(synchronize_session=False)
        Grievance.query.filter_by(student_id=student.id).delete()
        Certificate.query.filter_by(student_id=student.id).delete()
        LeaveApplication.query.filter_by(student_id=student.id).delete()
        db.session.flush()

        user = User.query.get(student.user_id)
        db.session.delete(student)
        db.session.flush()
        if user:
            db.session.delete(user)
        deleted += 1

    db.session.commit()
    flash(f'\U0001f5d1\ufe0f {deleted} student(s) deleted successfully.', 'success')
    return redirect(request.referrer or url_for('admin.students'))


# ── Bulk Move Students to Another Class ──────────────────────────────────────
@admin_bp.route('/students/bulk-move-class', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER)
def bulk_move_class():
    student_ids_str = request.form.get('student_ids', '')
    new_class_id    = request.form.get('new_class_id', type=int)

    if not student_ids_str or not new_class_id:
        flash('Please select students and a target class.', 'warning')
        return redirect(request.referrer or url_for('admin.students'))

    # Validate target class is within scope
    target_class = Class.query.get(new_class_id)
    if not target_class:
        flash('Target class not found.', 'danger')
        return redirect(request.referrer or url_for('admin.students'))

    if current_user.role == Roles.HOD:
        dept_id = _hod_dept_id()
        if target_class.department_id != dept_id:
            flash('You can only move students within your department.', 'danger')
            return redirect(request.referrer or url_for('admin.students'))
    elif current_user.role == Roles.CLASS_TEACHER:
        cls = _ct_class()
        if not cls or target_class.id != cls.id:
            flash('You can only assign students to your own class.', 'danger')
            return redirect(request.referrer or url_for('admin.students'))

    student_ids = [int(s.strip()) for s in student_ids_str.split(',') if s.strip().isdigit()]
    students = Student.query.filter(Student.id.in_(student_ids)).all()
    moved = 0
    new_year     = target_class.year if target_class.year else 1
    new_semester = (new_year * 2) - 1

    for student in students:
        # Scope check on source class
        if current_user.role == Roles.HOD:
            dept_id = _hod_dept_id()
            src_cls = Class.query.get(student.class_id)
            if not src_cls or src_cls.department_id != dept_id:
                continue
        elif current_user.role == Roles.CLASS_TEACHER:
            src_cls = _ct_class()
            if not src_cls or student.class_id != src_cls.id:
                continue

        student.class_id     = new_class_id
        student.current_year = new_year
        student.semester     = new_semester
        moved += 1

    db.session.commit()
    flash(f'📦 {moved} student(s) moved to {target_class.name}.', 'success')
    return redirect(request.referrer or url_for('admin.students'))


# ── Bulk Approve Students ─────────────────────────────────────────────────────
@admin_bp.route('/students/bulk-approve', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER)
def bulk_approve_students():
    student_ids_str = request.form.get('student_ids', '')
    if not student_ids_str:
        flash('No students selected.', 'warning')
        return redirect(request.referrer or url_for('admin.students'))

    student_ids = [int(s.strip()) for s in student_ids_str.split(',') if s.strip().isdigit()]
    students = Student.query.filter(Student.id.in_(student_ids)).all()
    approved = 0
    for student in students:
        if current_user.role == Roles.HOD:
            dept_id = _hod_dept_id()
            cls = Class.query.get(student.class_id)
            if not cls or cls.department_id != dept_id:
                continue
        elif current_user.role == Roles.CLASS_TEACHER:
            cls = _ct_class()
            if not cls or student.class_id != cls.id:
                continue

        if student.approval_status != ApprovalStatus.APPROVED:
            student.approval_status = ApprovalStatus.APPROVED
            student.approved_by     = current_user.id
            student.user.status     = Status.ACTIVE
            approved += 1

    db.session.commit()
    flash(f'✅ {approved} student(s) approved.', 'success')
    return redirect(request.referrer or url_for('admin.students'))


# ── Export Selected Students to Excel ────────────────────────────────────────
@admin_bp.route('/students/export-excel', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER, Roles.TEACHER)
def export_students_excel():
    student_ids_str = request.form.get('student_ids', '')
    if not student_ids_str:
        flash('No students selected for export.', 'warning')
        return redirect(request.referrer or url_for('admin.students'))

    student_ids = [int(s.strip()) for s in student_ids_str.split(',') if s.strip().isdigit()]
    # Scope: only export students this user is allowed to see
    students_all = Student.query.filter(Student.id.in_(student_ids)).all()
    students = []
    for s in students_all:
        if current_user.role == Roles.HOD:
            dept_id = _hod_dept_id()
            cls = Class.query.get(s.class_id)
            if cls and cls.department_id == dept_id:
                students.append(s)
        elif current_user.role == Roles.CLASS_TEACHER:
            cls = _ct_class()
            if cls and s.class_id == cls.id:
                students.append(s)
        elif current_user.role == Roles.TEACHER:
            taught_class_ids = [
                sub.class_id for sub in Subject.query.filter_by(teacher_id=current_user.id).all()
                if sub.class_id
            ]
            if s.class_id in taught_class_ids or s.tg_id == current_user.id:
                students.append(s)
        else:  # SUPER_ADMIN
            students.append(s)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students Export'

    from openpyxl.styles import Font, PatternFill, Alignment
    headers = ['Roll No', 'PRN', 'Name', 'Email', 'Phone',
               'Class', 'Gender', 'Category', 'Blood Group',
               'DOB', 'Mother Name', 'TG Name', 'Status']
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', fgColor='1A3C5E')
    header_align = Alignment(horizontal='center', vertical='center')

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = max(15, len(h) + 4)

    for r_idx, s in enumerate(students, start=2):
        cls = Class.query.get(s.class_id)
        tg  = User.query.get(s.tg_id) if s.tg_id else None
        ws.append([
            s.roll_no or '',
            s.prn or '',
            s.user.name,
            s.user.email,
            s.user.phone or '',
            cls.name if cls else '',
            'Male' if s.gender == 'M' else ('Female' if s.gender == 'F' else ''),
            s.category or '',
            s.blood_group or '',
            str(s.dob) if s.dob else '',
            s.mother_name or '',
            tg.name if tg else '',
            s.approval_status or '',
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     download_name='students_export.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@admin_bp.route('/departments')
@login_required
@role_required(Roles.SUPER_ADMIN)
def departments():
    depts    = Department.query.all()
    teachers = User.query.filter(User.role.in_([Roles.HOD, Roles.TEACHER, Roles.CLASS_TEACHER])).all()
    return render_template('admin/departments.html', departments=depts, teachers=teachers)


@admin_bp.route('/departments/create', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN)
def create_department():
    name   = request.form.get('name', '').strip()
    code   = request.form.get('code', '').strip()
    hod_id = request.form.get('hod_id') or None
    if name:
        dept = Department(name=name, code=code, hod_id=hod_id)
        db.session.add(dept)
        if hod_id:
            user = User.query.get(hod_id)
            if user and user.role != Roles.SUPER_ADMIN:
                user.role = Roles.HOD
        db.session.commit()
        flash(f'Department "{name}" created.', 'success')
    return redirect(url_for('admin.departments'))


@admin_bp.route('/departments/edit/<int:id>', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN)
def edit_department(id):
    dept = Department.query.get_or_404(id)
    dept.name = request.form.get('name', dept.name).strip()
    dept.code = request.form.get('code', dept.code).strip()
    new_hod_id = request.form.get('hod_id') or None

    if str(new_hod_id) != str(dept.hod_id):
        dept.hod_id = new_hod_id
        if new_hod_id:
            user = User.query.get(new_hod_id)
            if user and user.role != Roles.SUPER_ADMIN:
                user.role = Roles.HOD

    db.session.commit()
    flash(f'Department "{dept.name}" updated.', 'success')
    return redirect(url_for('admin.departments'))


# ── Classes ──────────────────────────────────────────────────────────────────
@admin_bp.route('/classes')
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD)
def classes():
    if current_user.role == Roles.HOD:
        dept_id = _hod_dept_id()
        classes_list = Class.query.filter_by(department_id=dept_id).all() if dept_id else []
        departments  = Department.query.filter_by(id=dept_id).all() if dept_id else []
    else:
        classes_list = Class.query.all()
        departments  = Department.query.all()

    # Teachers eligible to be CT (includes TEACHER role so HOD can promote them)
    teachers = User.query.filter(User.role.in_(
        [Roles.CLASS_TEACHER, Roles.TEACHER, Roles.HOD]
    )).all()
    return render_template('admin/classes.html', classes=classes_list,
                           departments=departments, teachers=teachers)


@admin_bp.route('/classes/create', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD)
def create_class():
    name    = request.form.get('name', '').strip()
    year    = request.form.get('year', 1)
    section = request.form.get('section', 'A')
    dept_id = request.form.get('department_id')
    ct_id   = request.form.get('class_teacher_id') or None

    # HOD can only create classes in their own department
    if current_user.role == Roles.HOD:
        dept_id = _hod_dept_id()

    if name:
        cls = Class(name=name, year=year, section=section,
                    department_id=dept_id, class_teacher_id=ct_id)
        db.session.add(cls)
        # Promote Class Teacher role
        if ct_id:
            ct_user = User.query.get(ct_id)
            if ct_user and ct_user.role == Roles.TEACHER:
                ct_user.role = Roles.CLASS_TEACHER
        db.session.commit()
        flash(f'Class "{name}" created.', 'success')
    return redirect(url_for('admin.classes'))


@admin_bp.route('/classes/edit/<int:id>', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD)
def edit_class(id):
    cls = Class.query.get_or_404(id)

    # HOD can only edit classes in their own department
    if current_user.role == Roles.HOD:
        dept_id = _hod_dept_id()
        if cls.department_id != dept_id:
            abort(403)

    cls.name    = request.form.get('name', cls.name).strip()
    cls.year    = request.form.get('year', cls.year)
    cls.section = request.form.get('section', cls.section)
    new_ct_id   = request.form.get('class_teacher_id') or None
    cls.class_teacher_id = new_ct_id

    # Promote new Class Teacher
    if new_ct_id:
        ct_user = User.query.get(new_ct_id)
        if ct_user and ct_user.role == Roles.TEACHER:
            ct_user.role = Roles.CLASS_TEACHER

    db.session.commit()
    flash(f'Class "{cls.name}" updated.', 'success')
    return redirect(url_for('admin.classes'))


# ── Subjects ──────────────────────────────────────────────────────────────────
@admin_bp.route('/subjects')
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER)
def subjects():
    if current_user.role == Roles.HOD:
        dept_id  = _hod_dept_id()
        class_ids = [c.id for c in Class.query.filter_by(department_id=dept_id).all()] if dept_id else []
        subjects_list = Subject.query.filter(Subject.class_id.in_(class_ids)).all()
        classes_list  = Class.query.filter_by(department_id=dept_id).all() if dept_id else []
    elif current_user.role == Roles.CLASS_TEACHER:
        cls = _ct_class()
        subjects_list = Subject.query.filter_by(class_id=cls.id).all() if cls else []
        classes_list  = [cls] if cls else []
    else:
        subjects_list = Subject.query.all()
        classes_list  = Class.query.all()

    teachers = User.query.filter(User.role.in_(
        [Roles.TEACHER, Roles.CLASS_TEACHER, Roles.HOD]
    )).all()
    return render_template('admin/subjects.html', subjects=subjects_list,
                           classes=classes_list, teachers=teachers)


@admin_bp.route('/subjects/create', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER)
def create_subject():
    name       = request.form.get('name', '').strip()
    code       = request.form.get('code', '').strip()
    class_id   = request.form.get('class_id')
    teacher_id = request.form.get('teacher_id') or None
    credits    = int(request.form.get('credits', 3))
    is_elective = request.form.get('is_elective') == 'on'

    # Scope check
    if current_user.role == Roles.CLASS_TEACHER:
        cls = _ct_class()
        if not cls or str(cls.id) != str(class_id):
            abort(403)
    elif current_user.role == Roles.HOD:
        dept_id = _hod_dept_id()
        cls = Class.query.get(class_id)
        if not cls or cls.department_id != dept_id:
            abort(403)

    if name:
        sub = Subject(name=name, code=code, class_id=class_id,
                      teacher_id=teacher_id, credits=credits, is_elective=is_elective)
        db.session.add(sub)
        db.session.commit()
        flash(f'Subject "{name}" created.', 'success')
    return redirect(url_for('admin.subjects'))


@admin_bp.route('/subjects/edit/<int:id>', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER)
def edit_subject(id):
    sub = Subject.query.get_or_404(id)

    # Scope check
    if current_user.role == Roles.CLASS_TEACHER:
        cls = _ct_class()
        if not cls or sub.class_id != cls.id:
            abort(403)
    elif current_user.role == Roles.HOD:
        dept_id = _hod_dept_id()
        cls = Class.query.get(sub.class_id)
        if not cls or cls.department_id != dept_id:
            abort(403)

    sub.name       = request.form.get('name', sub.name).strip()
    sub.code       = request.form.get('code', sub.code).strip()
    new_class_id   = request.form.get('class_id', sub.class_id)
    # Validate the NEW class_id is still within scope
    if str(new_class_id) != str(sub.class_id):
        if current_user.role == Roles.CLASS_TEACHER:
            cls = _ct_class()
            if not cls or str(cls.id) != str(new_class_id):
                abort(403)
        elif current_user.role == Roles.HOD:
            dept_id = _hod_dept_id()
            target_cls = Class.query.get(new_class_id)
            if not target_cls or target_cls.department_id != dept_id:
                abort(403)
    sub.class_id   = new_class_id
    sub.teacher_id = request.form.get('teacher_id') or None
    sub.credits    = int(request.form.get('credits', sub.credits))
    sub.is_elective = request.form.get('is_elective') == 'on'
    db.session.commit()
    flash(f'Subject "{sub.name}" updated.', 'success')
    return redirect(url_for('admin.subjects'))


@admin_bp.route('/subjects/<int:id>/enroll', methods=['GET', 'POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER)
def enroll_elective(id):
    sub = Subject.query.get_or_404(id)
    if not sub.is_elective:
        flash('This is a regular subject. All class students are automatically enrolled.', 'info')
        return redirect(url_for('admin.subjects'))

    # Scope check
    if current_user.role == Roles.CLASS_TEACHER:
        cls = _ct_class()
        if not cls or sub.class_id != cls.id:
            abort(403)
    elif current_user.role == Roles.HOD:
        dept_id = _hod_dept_id()
        cls = Class.query.get(sub.class_id)
        if not cls or cls.department_id != dept_id:
            abort(403)

    from app.models import Student, ApprovalStatus
    # Get all approved students in this class
    all_students = Student.query.filter_by(class_id=sub.class_id, approval_status=ApprovalStatus.APPROVED).order_by(Student.roll_no).all()

    if request.method == 'POST':
        selected_ids = request.form.getlist('student_ids')
        selected_students = Student.query.filter(Student.id.in_(selected_ids)).all() if selected_ids else []
        sub.enrolled_students = selected_students
        db.session.commit()
        flash(f'Enrollment updated for {sub.name}. {len(selected_students)} students enrolled.', 'success')
        return redirect(url_for('admin.subjects'))

    return render_template('admin/subject_enroll.html', subject=sub, students=all_students)



# ── Audit Logs ───────────────────────────────────────────────────────────────
@admin_bp.route('/audit-logs')
@login_required
@role_required(Roles.SUPER_ADMIN)
def audit_logs():
    from app.models import AuditLog
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(200).all()
    return render_template('admin/audit_logs.html', logs=logs)


# ── Student List ──────────────────────────────────────────────────────────────
@admin_bp.route('/students')
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER, Roles.TEACHER)
def students():
    search        = request.args.get('search', '').strip()
    category      = request.args.get('category', '')
    gender        = request.args.get('gender', '')
    blood         = request.args.get('blood', '')
    dept_filter   = request.args.get('department_id', type=int)
    class_filter  = request.args.get('class_id', type=int)
    year_filter   = request.args.get('year', type=int)
    page          = request.args.get('page', 1, type=int)

    query = Student.query.join(User, Student.user_id == User.id)

    available_depts   = []
    available_classes = []

    # ── Row-Level Access Controls & Dropdown Data ──────────────────────────
    if current_user.role == Roles.SUPER_ADMIN:
        available_depts = Department.query.all()
        if dept_filter:
            available_classes = Class.query.filter_by(department_id=dept_filter).all()
            query = query.join(Class, Student.class_id == Class.id).filter(Class.department_id == dept_filter)
            if class_filter:
                query = query.filter(Student.class_id == class_filter)
        elif class_filter:
            available_classes = Class.query.all()
            query = query.filter(Student.class_id == class_filter)
        else:
            available_classes = Class.query.all()

    elif current_user.role == Roles.HOD:
        dept_id = _hod_dept_id()
        dept_filter = dept_id  # Force lock dept
        available_classes = Class.query.filter_by(department_id=dept_id).all()
        if dept_id:
            query = query.join(Class, Student.class_id == Class.id).filter(Class.department_id == dept_id)
            if class_filter:
                query = query.filter(Student.class_id == class_filter)
        else:
            query = query.filter(Student.id == 0)

    elif current_user.role == Roles.CLASS_TEACHER:
        cls = _ct_class()
        if cls:
            class_filter = cls.id  # Force lock class
            query = query.filter(Student.class_id == cls.id)
        else:
            query = query.filter(Student.id == 0)

    elif current_user.role == Roles.TEACHER:
        # Collect class_ids from subjects this teacher is assigned to
        taught_class_ids = [
            s.class_id for s in Subject.query.filter_by(teacher_id=current_user.id).all()
            if s.class_id
        ]
        # Also include classes where they are TG
        tg_student_ids = [s.id for s in Student.query.filter_by(tg_id=current_user.id).all()]

        if taught_class_ids or tg_student_ids:
            query = query.filter(
                db.or_(
                    Student.class_id.in_(taught_class_ids),
                    Student.id.in_(tg_student_ids)
                )
            )
        else:
            query = query.filter(Student.id == 0)

    # ── Common Filters ─────────────────────────────────────────────────────
    if year_filter:
        query = query.filter(Student.current_year == year_filter)
    if search:
        query = query.filter(
            db.or_(
                User.name.ilike(f'%{search}%'),
                Student.prn.ilike(f'%{search}%'),
                Student.roll_no.ilike(f'%{search}%'),
            )
        )
    if category:
        query = query.filter(Student.category == category)
    if gender:
        query = query.filter(Student.gender == gender)
    if blood:
        query = query.filter(Student.blood_group == blood)

    pagination    = query.order_by(Student.roll_no).paginate(page=page, per_page=20, error_out=False)
    students_list = pagination.items

    total     = query.count()
    male_ct   = query.filter(Student.gender == 'M').count()
    female_ct = query.filter(Student.gender == 'F').count()

    categories   = db.session.query(Student.category).filter(
        Student.category != None).distinct().order_by(Student.category).all()
    blood_groups = db.session.query(Student.blood_group).filter(
        Student.blood_group != None).distinct().order_by(Student.blood_group).all()

    # Teachers for TG assignment dropdown
    teachers = User.query.filter(
        User.role.in_([Roles.TEACHER, Roles.CLASS_TEACHER, Roles.HOD])
    ).all()

    # ── Upload-modal class list: always the full scope, never affected by current filters ──
    if current_user.role == Roles.SUPER_ADMIN:
        upload_classes = Class.query.order_by(Class.name).all()
    elif current_user.role == Roles.HOD:
        _hd = _hod_dept_id()
        upload_classes = Class.query.filter_by(department_id=_hd).order_by(Class.name).all() if _hd else []
    elif current_user.role == Roles.CLASS_TEACHER:
        _cc = _ct_class()
        upload_classes = [_cc] if _cc else []
    else:
        upload_classes = []

    return render_template('admin/students.html',
        students=students_list, pagination=pagination,
        search=search, category=category, gender=gender, blood=blood,
        dept_filter=dept_filter, class_filter=class_filter, year_filter=year_filter,
        available_depts=available_depts, available_classes=available_classes,
        upload_classes=upload_classes,
        total=total, male_ct=male_ct, female_ct=female_ct,
        categories=[c[0] for c in categories],
        blood_groups=[b[0] for b in blood_groups],
        teachers=teachers,
    )





@admin_bp.route('/students/<int:id>')
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER, Roles.TEACHER)
def student_detail(id):
    student = Student.query.get_or_404(id)

    # Scope check
    if current_user.role == Roles.HOD:
        dept_id = _hod_dept_id()
        cls = Class.query.get(student.class_id)
        if not cls or cls.department_id != dept_id:
            abort(403)
    elif current_user.role == Roles.CLASS_TEACHER:
        cls = _ct_class()
        if not cls or student.class_id != cls.id:
            abort(403)
    elif current_user.role == Roles.TEACHER:
        taught_class_ids = [
            s.class_id for s in Subject.query.filter_by(teacher_id=current_user.id).all()
            if s.class_id
        ]
        is_tg = (student.tg_id == current_user.id)
        in_class = (student.class_id in taught_class_ids)
        if not is_tg and not in_class:
            abort(403)

    teachers = User.query.filter(
        User.role.in_([Roles.TEACHER, Roles.CLASS_TEACHER, Roles.HOD])
    ).all()
    return render_template('admin/student_detail.html', student=student, teachers=teachers)


@admin_bp.route('/students/<int:id>/update-batch', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER)
def update_batch(id):
    student = Student.query.get_or_404(id)

    # Scope check
    if current_user.role == Roles.HOD:
        dept_id = _hod_dept_id()
        cls = Class.query.get(student.class_id)
        if not cls or cls.department_id != dept_id:
            abort(403)
    elif current_user.role == Roles.CLASS_TEACHER:
        cls = _ct_class()
        if not cls or student.class_id != cls.id:
            abort(403)

    new_batch = request.form.get('batch', '').strip().upper() or None
    if new_batch and new_batch not in ['S1', 'S2', 'S3']:
        flash('Invalid batch. Choose S1, S2, or S3.', 'danger')
        return redirect(url_for('admin.student_detail', id=id))

    student.batch = new_batch
    db.session.commit()

    batch_label = f'Batch {new_batch}' if new_batch else 'No batch'
    flash(f'{student.user.name} assigned to {batch_label}.', 'success')
    return redirect(url_for('admin.student_detail', id=id) + '#tab-academic')


# ── Add Teacher (Manual) ─────────────────────────────────────────────────────
@admin_bp.route('/teachers/add', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD)
def add_teacher():
    name        = request.form.get('name', '').strip()
    email       = request.form.get('email', '').strip().lower()
    phone       = request.form.get('phone', '').strip()
    password    = request.form.get('password', 'csmss@123').strip() or 'csmss@123'
    role        = request.form.get('role', Roles.TEACHER)
    dept_id     = request.form.get('department_id') or None
    designation = request.form.get('designation', '').strip()

    # HOD can only add to their own department
    if current_user.role == Roles.HOD:
        dept_id = str(_hod_dept_id())

    if not name or not email:
        flash('Name and Email are required.', 'danger')
        return redirect(url_for('admin.users'))

    if User.query.filter_by(email=email).first():
        flash(f'Email {email} is already registered.', 'danger')
        return redirect(url_for('admin.users'))

    if role not in Roles.ALL:
        role = Roles.TEACHER
    # Privilege escalation guard: HOD cannot create HOD/SUPER_ADMIN users
    if current_user.role == Roles.HOD and role in [Roles.SUPER_ADMIN, Roles.HOD]:
        role = Roles.TEACHER

    pw_hash = bcrypt.generate_password_hash(password).decode('utf-8')
    user = User(name=name, email=email, phone=phone, password_hash=pw_hash,
                role=role, status=Status.ACTIVE, must_change_password=True)
    db.session.add(user)
    db.session.flush()

    teacher = Teacher(user_id=user.id, department_id=dept_id,
                      designation=designation, teacher_type='subject')
    db.session.add(teacher)
    db.session.commit()
    flash(f'Teacher "{name}" added successfully! Default password: csmss@123', 'success')
    return redirect(url_for('admin.users'))


# ── Add Student (Manual) ─────────────────────────────────────────────────────
@admin_bp.route('/students/add', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER)
def add_student():
    name        = request.form.get('name', '').strip()
    email       = request.form.get('email', '').strip().lower()
    phone       = request.form.get('phone', '').strip()
    password    = request.form.get('password', 'csmss@123').strip() or 'csmss@123'
    prn         = request.form.get('prn', '').strip()
    roll_no     = request.form.get('roll_no', '').strip()
    class_id    = request.form.get('class_id') or None
    dob         = request.form.get('dob', '').strip()
    gender      = request.form.get('gender', '').strip()
    category    = request.form.get('category', '').strip()
    mother_name = request.form.get('mother_name', '').strip()

    # Class Teacher: lock to their class
    if current_user.role == Roles.CLASS_TEACHER:
        cls = _ct_class()
        class_id = cls.id if cls else None
    elif current_user.role == Roles.HOD:
        # HOD can only add students to classes in their own department
        dept_id = _hod_dept_id()
        if class_id:
            target_cls = Class.query.get(int(class_id))
            if not target_cls or target_cls.department_id != dept_id:
                flash('You can only add students to your department classes.', 'danger')
                return redirect(url_for('admin.students'))

    if not name or not email:
        flash('Name and Email are required.', 'danger')
        return redirect(url_for('admin.students'))

    if User.query.filter_by(email=email).first():
        flash(f'Email {email} is already registered.', 'danger')
        return redirect(url_for('admin.students'))

    pw_hash = bcrypt.generate_password_hash(password).decode('utf-8')
    user = User(name=name, email=email, phone=phone or None,
                password_hash=pw_hash, role=Roles.STUDENT,
                status=Status.ACTIVE, must_change_password=True)
    db.session.add(user)
    db.session.flush()

    student = Student(
        user_id=user.id,
        class_id=class_id,
        prn=prn or None,
        roll_no=roll_no or None,
        approval_status=ApprovalStatus.APPROVED,
        approved_by=current_user.id,
        dob=dob or None,
        gender=gender if gender in ['M', 'F'] else None,
        category=category or None,
        mother_name=mother_name or None,
    )
    db.session.add(student)
    db.session.commit()
    flash(f'Student "{name}" added successfully!', 'success')
    return redirect(url_for('admin.students'))


# ── Upload Students via Excel ────────────────────────────────────────────────
@admin_bp.route('/students/upload-excel', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER)
def upload_students_excel():
    file = request.files.get('excel_file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('Please upload a valid Excel file (.xlsx or .xls).', 'danger')
        return redirect(url_for('admin.students'))

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
    except Exception:
        flash('Could not read the Excel file. Please use the correct format.', 'danger')
        return redirect(url_for('admin.students'))

    added   = 0
    skipped = 0
    errors  = []

    # ── Determine scope & default class ──────────────────────────────────────
    hod_dept_id = _hod_dept_id() if current_user.role == Roles.HOD else None
    ct_class    = _ct_class()     if current_user.role == Roles.CLASS_TEACHER else None

    if current_user.role == Roles.CLASS_TEACHER:
        if not ct_class:
            flash('You are not assigned as a Class Teacher for any class.', 'danger')
            return redirect(url_for('admin.students'))
        form_class_id = ct_class.id
    else:
        form_class_id = request.form.get('class_id', type=int)

    if not form_class_id:
        flash('Please select a default class before uploading the Excel file.', 'danger')
        return redirect(url_for('admin.students'))

    default_class = Class.query.get(form_class_id)
    if not default_class:
        flash('Selected class not found. Please try again.', 'danger')
        return redirect(url_for('admin.students'))

    if current_user.role == Roles.HOD and hod_dept_id:
        if default_class.department_id != hod_dept_id:
            flash('Invalid class selected for your department.', 'danger')
            return redirect(url_for('admin.students'))

    # ── Build a scoped class lookup: {lowercase_name: class_obj} ────────────────
    # This lets each Excel row override class via its class_name column.
    # Only classes within the user's scope are included (security).
    if current_user.role == Roles.SUPER_ADMIN:
        scoped_classes = Class.query.all()
    elif current_user.role == Roles.HOD and hod_dept_id:
        scoped_classes = Class.query.filter_by(department_id=hod_dept_id).all()
    elif current_user.role == Roles.CLASS_TEACHER and ct_class:
        scoped_classes = [ct_class]
    else:
        scoped_classes = []

    # Multiple lookup keys per class for flexible matching:
    # e.g. "SE ECE", "se ece", "SE-ECE" all resolve to the same class
    class_lookup = {}
    for cls in scoped_classes:
        keys = [
            cls.name.strip().lower(),
            cls.name.strip().lower().replace(' ', '-'),
            cls.name.strip().lower().replace('-', ' '),
        ]
        for k in keys:
            class_lookup[k] = cls

    # ── Parse headers ────────────────────────────────────────────────────────
    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
    for req in ['name', 'email']:
        if req not in headers:
            flash(f'Excel is missing required column: "{req}". Download the format template.', 'danger')
            return redirect(url_for('admin.students'))

    def col(row, name):
        try:
            idx = headers.index(name)
            val = row[idx].value
            return str(val).strip() if val is not None else ''
        except (ValueError, IndexError):
            return ''

    # ── Performance: hash once, pre-fetch sets ──────────────────────────────────
    default_pw_hash = bcrypt.generate_password_hash('csmss@123').decode('utf-8')
    existing_emails = set(r[0] for r in db.session.query(User.email).all())
    existing_prns   = set(r[0] for r in db.session.query(Student.prn).filter(Student.prn != None).all())
    BATCH_SIZE      = 50

    # ── Process rows ───────────────────────────────────────────────────────────
    for row in ws.iter_rows(min_row=2):
        if all(cell.value is None for cell in row):
            continue

        name        = col(row, 'name')
        email       = col(row, 'email').lower()
        phone       = col(row, 'phone')
        prn         = col(row, 'prn')
        roll_no     = col(row, 'roll_no')
        class_name  = col(row, 'class_name')   # per-row class override
        dob         = col(row, 'dob')
        gender      = col(row, 'gender').upper()[:1]
        category    = col(row, 'category').upper()
        mother_name = col(row, 'mother_name')

        if not name or not email:
            skipped += 1
            continue

        if email in existing_emails:
            errors.append(f'Skipped (email exists): {email}')
            skipped += 1
            continue

        if prn and prn in existing_prns:
            errors.append(f'Skipped (PRN exists): {prn}')
            skipped += 1
            continue

        # ── Resolve class for THIS row ───────────────────────────────────────
        # Priority: class_name column > dropdown selection
        resolved = class_lookup.get(class_name.strip().lower()) if class_name else None
        if resolved is None:
            # Fall back to the dropdown default
            resolved = default_class

        row_class_id = resolved.id
        row_year     = resolved.year if resolved.year else 1
        row_semester = (row_year * 2) - 1

        user = User(name=name, email=email, phone=phone or None,
                    password_hash=default_pw_hash, role=Roles.STUDENT,
                    status=Status.ACTIVE, must_change_password=True)
        db.session.add(user)
        db.session.flush()  # get user.id for FK

        student = Student(user_id=user.id, class_id=row_class_id,
                          prn=prn or None, roll_no=roll_no or None,
                          approval_status=ApprovalStatus.APPROVED,
                          approved_by=current_user.id,
                          dob=dob or None,
                          gender=gender if gender in ['M', 'F'] else None,
                          category=category or None,
                          mother_name=mother_name or None,
                          current_year=row_year,
                          semester=row_semester)
        db.session.add(student)
        existing_emails.add(email)
        if prn:
            existing_prns.add(prn)
        added += 1

        if added % BATCH_SIZE == 0:
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                flash(f'Database error at row ~{added}: {str(e)}', 'danger')
                return redirect(url_for('admin.students'))

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Database error during upload: {str(e)}', 'danger')
        return redirect(url_for('admin.students'))

    msg = f'✅ Excel upload complete: {added} student(s) added, {skipped} skipped.'
    if errors:
        msg += f' Issues: {"; ".join(errors[:5])}'
    flash(msg, 'success' if added > 0 else 'warning')
    return redirect(url_for('admin.students'))


# ── Upload Teachers via Excel ────────────────────────────────────────────────
@admin_bp.route('/teachers/upload-excel', methods=['POST'])
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD)
def upload_teachers_excel():
    file = request.files.get('excel_file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('Please upload a valid Excel file (.xlsx or .xls).', 'danger')
        return redirect(url_for('admin.users'))

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
    except Exception:
        flash('Could not read the Excel file. Please use the correct format.', 'danger')
        return redirect(url_for('admin.users'))

    added = 0
    skipped = 0
    errors = []

    hod_dept_id = _hod_dept_id() if current_user.role == Roles.HOD else None

    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
    required = ['name', 'email']
    for req in required:
        if req not in headers:
            flash(f'Excel is missing required column: "{req}". Download the format template.', 'danger')
            return redirect(url_for('admin.users'))

    def col(row, name):
        try:
            idx = headers.index(name)
            val = row[idx].value
            return str(val).strip() if val is not None else ''
        except (ValueError, IndexError):
            return ''

    # ── Hash password once outside loop (same fix as student upload) ──────────
    default_pw_hash  = bcrypt.generate_password_hash('csmss@123').decode('utf-8')
    existing_emails  = set(r[0] for r in db.session.query(User.email).all())
    BATCH_SIZE       = 50

    for row in ws.iter_rows(min_row=2):
        if all(cell.value is None for cell in row):
            continue

        name        = col(row, 'name')
        email       = col(row, 'email').lower()
        phone       = col(row, 'phone')
        dept_code   = col(row, 'department_code').upper()
        designation = col(row, 'designation')
        role        = col(row, 'role').upper()

        if not name or not email:
            skipped += 1
            continue

        if email in existing_emails:
            errors.append(f'Skipped (email exists): {email}')
            skipped += 1
            continue

        if role not in Roles.ALL:
            role = Roles.TEACHER
        # Privilege escalation guard: HOD cannot create HOD/SUPER_ADMIN via Excel
        if current_user.role == Roles.HOD and role in [Roles.SUPER_ADMIN, Roles.HOD]:
            role = Roles.TEACHER

        # Resolve department
        dept_id = None
        if current_user.role == Roles.HOD:
            dept_id = hod_dept_id
        elif dept_code:
            dept_obj = Department.query.filter(Department.code.ilike(dept_code)).first()
            if dept_obj:
                dept_id = dept_obj.id

        user = User(name=name, email=email, phone=phone or None,
                    password_hash=default_pw_hash, role=role,
                    status=Status.ACTIVE, must_change_password=True)
        db.session.add(user)
        db.session.flush()

        teacher = Teacher(user_id=user.id, department_id=dept_id,
                          designation=designation or None, teacher_type='subject')
        db.session.add(teacher)
        existing_emails.add(email)
        added += 1

        if added % BATCH_SIZE == 0:
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                flash(f'Database error at row ~{added}: {str(e)}', 'danger')
                return redirect(url_for('admin.users'))

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Database error during upload: {str(e)}', 'danger')
        return redirect(url_for('admin.users'))

    msg = f'✅ Excel upload complete: {added} teacher(s) added, {skipped} skipped.'
    if errors:
        msg += f' Issues: {"; ".join(errors[:5])}'
    flash(msg, 'success' if added > 0 else 'warning')
    return redirect(url_for('admin.users'))


# ── Download Excel Format Templates ─────────────────────────────────────────
def _make_template_wb(headers, sample_row, title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', fgColor='1A3C5E')
    header_align = Alignment(horizontal='center', vertical='center')
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = max(18, len(h) + 4)
    for col_idx, val in enumerate(sample_row, 1):
        ws.cell(row=2, column=col_idx, value=val)
    return wb

@admin_bp.route('/students/download-template')
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD, Roles.CLASS_TEACHER)
def download_students_template():
    # ── Build scoped class list for the sample rows ─────────────────────────
    if current_user.role == Roles.SUPER_ADMIN:
        all_classes = Class.query.order_by(Class.name).limit(5).all()
    elif current_user.role == Roles.HOD:
        dept_id = _hod_dept_id()
        all_classes = Class.query.filter_by(department_id=dept_id).order_by(Class.name).all() if dept_id else []
    elif current_user.role == Roles.CLASS_TEACHER:
        cls = _ct_class()
        all_classes = [cls] if cls else []
    else:
        all_classes = []

    headers = ['name', 'email', 'phone', 'prn', 'roll_no', 'class_name',
               'dob', 'gender', 'category', 'mother_name']

    # Generate sample rows using REAL class names so users copy them correctly
    sample_rows = []
    for i, cls in enumerate(all_classes[:3], start=1):
        sample_rows.append([
            f'Student Name {i}',
            f'student{i}@csmss.edu',
            f'98765432{i:02d}',
            f'PRN2024{i:03d}',
            f'{i:02d}',
            cls.name,           # exact class name from DB
            '01-01-2004',
            'M' if i % 2 == 1 else 'F',
            'OPEN',
            f'Mother Name {i}',
        ])
    if not sample_rows:
        sample_rows = [['John Doe', 'john@csmss.edu', '9876543210',
                        'PRN2024001', '01', 'SE ECE',
                        '01-01-2004', 'M', 'OPEN', 'Jane Doe']]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', fgColor='1A3C5E')
    note_fill    = PatternFill('solid', fgColor='FFF2CC')   # yellow for class_name col
    header_align = Alignment(horizontal='center', vertical='center')

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = header_font
        cell.alignment = header_align
        # Highlight class_name column in yellow so users notice it
        cell.fill = note_fill if h == 'class_name' else header_fill
        ws.column_dimensions[cell.column_letter].width = max(18, len(h) + 4)

    for r_idx, sample in enumerate(sample_rows, start=2):
        for c_idx, val in enumerate(sample, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Add a note sheet listing all available class names
    ws_note = wb.create_sheet(title='Available Classes')
    ws_note.cell(row=1, column=1, value='Available class_name values (copy exactly into main sheet):')
    ws_note.cell(row=1, column=1).font = Font(bold=True)
    ws_note.column_dimensions['A'].width = 55
    for i, cls in enumerate(all_classes, start=2):
        ws_note.cell(row=i, column=1, value=cls.name)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='students_upload_format.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@admin_bp.route('/teachers/download-template')
@login_required
@role_required(Roles.SUPER_ADMIN, Roles.HOD)
def download_teachers_template():
    headers    = ['name', 'email', 'phone', 'department_code', 'designation', 'role']
    sample_row = ['Prof. Smith', 'smith@csmss.edu', '9876543211', 'CSE', 'Assistant Professor', 'TEACHER']
    wb = _make_template_wb(headers, sample_row, 'Teachers')
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='teachers_upload_format.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
